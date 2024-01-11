# -*- coding: utf-8 -*-

from datetime import timedelta, date, datetime
import base64

import os
from io import BytesIO
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment

from openpyxl.worksheet import filters
from odoo import api, fields, models



class SaleListReport(models.TransientModel):
    _name = 'sale.list.report'
    _description = 'Báo cáo tồn kho vật tư'

    year = fields.Selection(
        selection='years_selection',
        string="Năm",
        default=str(datetime.now().year), required=True)
    month = fields.Selection(
        selection='month_selection',
        string="Tháng",
        default=str(datetime.now().month), required=True)
    district_ids = fields.Many2many('res.country.state.district',string="Huyện",)


    def years_selection(self):
        y = datetime.now().year
        year_list = []
        while y != 1939:
            year_list.append((str(y), str(y)))
            y -= 1
        return year_list
    def month_selection(self):
        y = 12
        month_list = []
        while y != 0:
            month_list.append((str(y), str(y)))
            y -= 1
        return month_list


    def action_report(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        wb = openpyxl.load_workbook(dir_path + '%s..%stemplates%ssale_list_report.xlsx' % (os.sep, os.sep, os.sep))
        ws = wb['sale_list']
        current_year = self.year

        sql = '''select 	
                    sale_order.name as ma_bao_gia,
                    rp_1.customer_code as ma_khach_hang,
                    rp_1.id_number as cccd,
                    rp_1.name as khach_hang,
                    rp_1.phone as sdt,
                    zone_1.name as xom_kh,
                    ward_1.name as xa_kh,
                    district_1.name as huyen_kh,
                    state_1.name as tinh_kh,
                    country_1.name as quoc_gia_kh,
                    rp_2.contract_no as ma_hd,
                    qwaco_water_meter.name as ma_dhn,
                    zone_2.name as xom_dhn,
                    ward_2.name as xa_dhn,
                    district_2.name as huyen_dhn,
                    state_2.name as tinh_dhn,
                    country_2.name as quoc_gia_dhn,
                    product_pricelist.name as bang_gia,
                    sol_1.price_unit as đon_gia,
                    qwaco_water_meter_quantity_history.old_quantity as so_cu,
                    qwaco_water_meter_quantity_history.new_quantity as so_moi,
                    sol_1.product_uom_qty as slg_tieu_thu,
                    sol_2.product_uom_qty as giam_gia,
                    sol_1.product_uom_qty + sol_2.product_uom_qty as slg_thanh_toan,
                    sale_order.amount_untaxed as tien_truoc_thue,
                    sale_order.amount_tax as thue,
                    sale_order.amount_total as tien_sau_thue,
                    sale_order.is_paid as thanh_toan,
                    DATE(sale_order.paid_date) as ngay_thanh_toan
            from sale_order
            left join res_partner as rp_1 on sale_order.partner_id = rp_1.id
            left join res_country_zone as zone_1 on rp_1.zone_id = zone_1.id
            left join res_country_ward as ward_1 on rp_1.ward_id = ward_1.id
            left join res_country_state_district as district_1 on rp_1.district_id = district_1.id
            left join res_country_state as state_1 on rp_1.state_id = state_1.id
            left join res_country as country_1 on rp_1.country_id = country_1.id
            left join qwaco_water_meter on sale_order.water_meter_id = qwaco_water_meter.id
            left join res_partner as rp_2 on qwaco_water_meter.partner_id = rp_2.id
            left join res_country_zone as zone_2 on rp_2.zone_id = zone_2.id
            left join res_country_ward as ward_2 on rp_2.ward_id = ward_2.id
            left join res_country_state_district as district_2 on rp_2.district_id = district_2.id
            left join res_country_state as state_2 on rp_2.state_id = state_2.id
            left join res_country as country_2 on rp_2.country_id = country_2.id
            left join qwaco_water_meter_quantity_history on sale_order.id = qwaco_water_meter_quantity_history.order_id 
            left join product_pricelist on sale_order.pricelist_id = product_pricelist.id
            left join
                (select sale_order_line.*,
                product_template.id as tmpl_id
                from sale_order_line
                left join product_product on sale_order_line.product_id = product_product.id
                left join product_template on product_product.product_tmpl_id = product_template.id
                ) as sol_1
            on sale_order.id = sol_1.order_id and sol_1.tmpl_id = 1
            left join
                (select sale_order_line.*,
                product_template.id as tmpl_id
                from sale_order_line
                left join product_product on sale_order_line.product_id = product_product.id
                left join product_template on product_product.product_tmpl_id = product_template.id
                ) as sol_2
            on sale_order.id = sol_2.order_id and sol_2.tmpl_id = 5
            where sale_order.state not in ('cancel')
                  and district_1.id in {huyen}
                  and extract ('year' from sale_order.date_order)::character varying ='{year}'
                    and	extract ('month' from sale_order.date_order)::character varying = '{month}'
            '''.format(month= self.month,year= current_year, huyen=tuple(self.district_ids.ids + [0, 0]))
        # print(sql)
        self._cr.execute(sql)
        recs = self._cr.dictfetchall()

        highlight = NamedStyle(name="highlight")
        bd1 = Side(style='thin', color="000000")
        bd2 = Side(style='dotted', color="000000")
        bd3 = Side(style='none')
        highlight.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd2)

        style_sum1 = NamedStyle(name="style_sum1")
        style_sum1.PatternFill = Font(color='0099CC00')
        style_sum1.number_format = '#,##0'
        style_sum1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd2)

        highlight1 = NamedStyle(name="highlight1")
        highlight1.font = Font(name='Verdana', size=12,bold=True)
        highlight1.border = Border(left=bd1, top=bd1, right=bd1, bottom=bd1)
        highlight1.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

        ws.cell(3, 6).value = "Tháng " + self.month+" năm "+self.year
        ws.cell(2, 21).value = "Tạo bởi:" + self.env.user.name
        ws.cell(3, 21).value = "Ngày tạo: " + str(datetime.now().date())

        row = 5

        for r in recs:
            ws.cell(row, 1).value, ws.cell(row, 1).style = r['ma_bao_gia'],highlight
            ws.cell(row, 2).value, ws.cell(row, 2).style = r['ma_khach_hang'],highlight
            ws.cell(row, 3).value, ws.cell(row, 3).style = r['cccd'],highlight
            ws.cell(row, 4).value, ws.cell(row, 4).style = r['khach_hang'],highlight
            ws.cell(row, 5).value, ws.cell(row, 5).style = r['sdt'],highlight
            ws.cell(row, 6).value, ws.cell(row, 6).style = '%s%s%s%s%s' % (r['xom_kh']['en_US']+',' if r.get('xom_kh') else ' ',
                                                    r.get('xa_kh')['en_US'] + ',' if r.get('xa_kh') else ' ',
                                                    r['huyen_kh']['en_US'] + ',' if r.get('huyen_kh') else ' ',
                                                    r.get('tinh_kh')+',' if r.get('tinh_kh') else ' ' ,
                                                    r['quoc_gia_kh']['vi_VN'] + ',' if r.get('quoc_gia_kh') else ' '),highlight
            ws.cell(row, 7).value, ws.cell(row, 7).style = r['ma_hd'],highlight
            ws.cell(row, 8).value, ws.cell(row, 8).style = r['ma_dhn'],highlight
            ws.cell(row, 9).value, ws.cell(row, 9).style = '%s%s%s%s%s' % (r['xom_dhn']['en_US']+',' if r.get('xom_dhn') else ' ',
                                                    r.get('xa_dhn')['en_US'] + ',' if r.get('xa_dhn') else ' ',
                                                    r['huyen_dhn']['en_US'] + ',' if r.get('huyen_dhn') else ' ',
                                                    r.get('tinh_dhn')+','  if r.get('tinh_dhn') else ' ' ,
                                                    r['quoc_gia_dhn']['vi_VN'] + ',' if r.get('quoc_gia_dhn') else ' '),highlight
            ws.cell(row, 10).value, ws.cell(row, 10).style = r['bang_gia']['en_US'],highlight
            ws.cell(row, 11).value, ws.cell(row, 11).style = r['đon_gia'],style_sum1
            ws.cell(row, 12).value , ws.cell(row, 12).style= r['so_cu'],style_sum1
            ws.cell(row, 13).value , ws.cell(row, 13).style= r['so_moi'],style_sum1
            ws.cell(row, 14).value , ws.cell(row, 14).style= r['slg_tieu_thu'],style_sum1
            ws.cell(row, 15).value , ws.cell(row, 15).style= r['giam_gia'],style_sum1
            ws.cell(row, 16).value, ws.cell(row, 16).style = r['slg_thanh_toan'],style_sum1
            ws.cell(row, 17).value , ws.cell(row, 17).style= r['tien_truoc_thue'],style_sum1
            ws.cell(row, 18).value , ws.cell(row, 18).style= r['thue'],style_sum1
            ws.cell(row, 19).value , ws.cell(row, 19).style= r['tien_sau_thue'],style_sum1
            ws.cell(row, 20).value , ws.cell(row, 20).style= 'Đã thanh toán' if r['thanh_toan'] == True else '' ,highlight
            ws.cell(row, 21).value, ws.cell(row, 21).style = str(r['ngay_thanh_toan']) ,highlight

            row += 1

        stream = BytesIO()
        wb.save(stream)
        xls = stream.getvalue()

        attachment_id = self.env['ir.attachment'].create({
            'name': 'Báo cáo tình trạng thanh toán.xlsx',
            'datas': base64.b64encode(xls),
            'type': 'binary',
        })
        # download
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/' + str(attachment_id.id) + '?download=true',
            'target': 'new',
        }
